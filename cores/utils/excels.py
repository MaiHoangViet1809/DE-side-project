from cores.utils.providers import ProvideBenchmark


def get_header_and_rows(excel_path: str, sample_size: int = 50):
    """
    Identify the first header row index and total rows containing data in an Excel file.

    Args:
        excel_path (str): Path to the Excel file.
        sample_size (int, optional): Number of rows to infer the header. Defaults to 50.

    Returns:
        tuple: A tuple containing:
            - int: Index of the header row.
            - int: Total rows below the header row.

    Examples:
        ```python
        header_row_index, total_rows = get_header_and_rows("example.xlsx", sample_size=50)
        print(header_row_index, total_rows)
        ```
    """
    import openpyxl
    wb = openpyxl.load_workbook(filename=excel_path,
                                read_only=False,
                                keep_vba=False,
                                data_only=True,
                                keep_links=False)
    first_sheet = wb.worksheets[0]
    sample_row = [row for row in first_sheet.iter_rows(values_only=True, min_row=1, max_row=sample_size)]
    sample_row_map = list(map(lambda x: len(set(x) - {None}), sample_row))
    max_non_null = max(sample_row_map)
    header_row_index = sample_row_map.index(max_non_null)
    actual_row = len(first_sheet['A']) - header_row_index - 1
    return header_row_index, actual_row


def read_excel(source_path, sample_size: int = 50, **kwargs):
    """
    Read an Excel file into a pandas DataFrame.

    Args:
        source_path (str): Path to the Excel file.
        sample_size (int, optional): Number of rows to infer header and data. Defaults to 50.
        kwargs: Additional arguments for pandas `read_excel`.

    Returns:
        pandas.DataFrame: The loaded data.

    Examples:
        ```python
        df = read_excel("example.xlsx", sample_size=50)
        print(df.head())
        ```
    """
    from pandas import read_excel as _read_excel_
    from typing import cast, Literal
    header_row_index, actual_row = get_header_and_rows(excel_path=source_path, sample_size=sample_size)
    engine = cast(Literal, kwargs.pop("engine", "openpyxl"))  # openpyxl
    df = _read_excel_(io=source_path,
                      header=header_row_index,
                      na_filter=False,
                      nrows=actual_row,
                      engine=engine,
                      **kwargs
                      )
    return df


@ProvideBenchmark  # noqa
def convert_xlsb_xlsx(file_name: str, data_address: str):
    """
    Convert an Excel binary file (.xlsb) to a standard Excel file (.xlsx).

    Args:
        file_name (str): Path to the .xlsb file.
        data_address (str): Data address (e.g., 'Sheet1!A1:Z99').

    Returns:
        str: Path to the converted .xlsx file.

    Examples:
        ```python
        new_file_path = convert_xlsb_xlsx("example.xlsb", "Sheet1!A1:Z99")
        print(new_file_path)
        ```
    """
    import pandas as pd
    from cores.utils.debug import print_table
    sheet_name = data_address.split("!")[0]
    sheet_name = sheet_name.removeprefix("'").removesuffix("'")

    if file_name.lower().endswith(".xlsb"):
        new_file = file_name.replace(".xlsb", ".xlsx")
        df = pd.read_excel(io=file_name, sheet_name=sheet_name, engine='pyxlsb')
        print(f"[convert_xlsb_xlsx] {sheet_name=} {file_name=} {new_file=}")
        print_table(df, 5)
        df.to_excel(excel_writer=new_file, sheet_name=sheet_name)
        return new_file
    else:
        raise ValueError(f"[convert_xlsb_xlsx] file: {file_name} is not support")


def validate_data_address(file_name: str, data_address: str):
    """
    Validate if the given data address refers to an existing sheet in an Excel file.

    Args:
        file_name (str): Path to the Excel file.
        data_address (str): Data address in A1 format (e.g., 'Sheet1!A1:Z99').

    Returns:
        bool: True if the sheet exists, False otherwise.

    Examples:
        ```python
        is_valid = validate_data_address("example.xlsx", "Sheet1!A1:Z99")
        print(is_valid)
        ```
    """
    from openpyxl import load_workbook
    list_sheets = load_workbook(file_name, read_only=True).sheetnames
    sheet_name = data_address.split("!")[0]
    sheet_name = sheet_name.removeprefix("'").removesuffix("'")
    return sheet_name in list_sheets


if __name__ == "__main__":
    ...
